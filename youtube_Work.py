from selenium import webdriver
import time
import sys, os
import mysql.connector
import datetime
import global_var
from openpyxl import load_workbook
from datetime import datetime


def test():
    data_count = 0
    # row_count = 3
    browser = webdriver.Chrome(executable_path='D:\\Youtube Work\\chromedriver.exe')
    browser.maximize_window()
    try:
        browser.switch_to.window(browser.window_handles[1])
        browser.close()
        browser.switch_to.window(browser.window_handles[0])
    except:
        browser.switch_to.window(browser.window_handles[0])

    # Search_keyword = 'heartfulness'
    browser.get('https://www.youtube.com/results?search_query='+str(global_var.Search_keyword))
    time.sleep(2)
    channel_link = global_var.channel_link_list
    now = datetime.now()  # current date and time
    date_time = now.strftime("%d-%m-%Y %H.%M.%S")
    book = load_workbook("D:\\Youtube Work\\Youtube Video Details.xlsx")
    # sheet_list = book.sheetnames
    # sheet = sheet_list[-1]
    # print(sheet_list)
    # sheet = book.get_sheet_by_name(str(sheet))
    # sheet.title = str(date_time)
    sheet = book.active
    max_row = sheet.max_row
    max_row += 1
    if max_row == 0:
        max_row = 3
    for cell in sheet['G']:
        global_var.video_links.append(cell.value)

    import wx
    app = wx.App()
    y = 1080
    a = True
    while a == True:
        try:
            dropdown_value = str(global_var.dropdown_value).partition('-')[2].strip()
            if len(global_var.video_links) < int(dropdown_value):
                for video_deatails in browser.find_elements_by_xpath('//*[@class="text-wrapper style-scope ytd-video-renderer"]'):
                    b = True
                    while b == True:
                        try:
                            video_deatails = video_deatails.get_attribute('outerHTML')
                            youtube_channel_link = video_deatails.partition('class="yt-simple-endpoint style-scope yt-formatted-string" spellcheck="false"')[2].partition(">")[0]
                            youtube_channel_link = youtube_channel_link.partition('href="')[2].partition('"')[0]
                            youtube_channel_link = "https://www.youtube.com"+youtube_channel_link.replace(' ', '')
                            # print(youtube_channel_link)

                            youtube_video_link = video_deatails.partition('yt-simple-endpoint style-scope ytd-video-renderer')[2].partition('>')[0]
                            youtube_video_link = youtube_video_link.partition('href="')[2].partition('"')[0]
                            youtube_video_link = "https://www.youtube.com" + youtube_video_link.replace(' ', '')
                            # print(youtube_video_link)

                            if youtube_video_link not in global_var.video_links:
                                global_var.video_links.append(youtube_video_link)
                                if youtube_channel_link not in channel_link:
                                    browser.execute_script("window.open('');")
                                    browser.switch_to.window(browser.window_handles[1])
                                    browser.get(str(youtube_video_link))
                                    time.sleep(3)
                                    video_Title = ''
                                    video_Views = ''
                                    video_Date = ''
                                    channel_subscribers = ''
                                    channel_name = ''
                                    desc = ''
                                    comments = ''
                                    hash_tag = ''

                                    keywordFound = False

                                    print('youtube video link: ',youtube_video_link)
                                    print('youtube channel link: ',youtube_channel_link)
                                    time.sleep(0.5)
                                    for video_Title in browser.find_elements_by_xpath('//*[@class="title style-scope ytd-video-primary-info-renderer"]'):
                                        video_Title = video_Title.get_attribute('innerText').strip()
                                        break
                                    print('Video Title: ', video_Title)
                                    time.sleep(0.5)
                                    for video_Views in browser.find_elements_by_xpath('//*[@class="view-count style-scope yt-view-count-renderer"]'):
                                        video_Views = video_Views.get_attribute('innerText').replace('views','').replace(',','').replace('watching','-LIVE').replace('watching now','-LIVE').strip().replace('now','')
                                        break
                                    print('Video Views: ', video_Views)
                                    time.sleep(0.5)
                                    for video_Date in browser.find_elements_by_xpath('//*[@id="date"]'):
                                        video_Date = video_Date.get_attribute('innerText').replace('•', '').strip()
                                        cont = False
                                        if 'hours' in video_Date:
                                            cont = True
                                        if 'min' in video_Date:
                                            cont = True
                                        if 'Hours' in video_Date:
                                            cont = True
                                        if 'Min' in video_Date:
                                            cont = True
                                        if 'Premiered' in video_Date:
                                            video_Date = video_Date.replace('Premiered','')
                                            video_Date = video_Date+' - Premiered'
                                        if 'Streamed live on' in video_Date:
                                            video_Date = video_Date.replace('Streamed live on','')
                                            video_Date = video_Date+' - Streamed'
                                        if cont ==  True:
                                            video_Date = now.strftime("%b %d, %Y")
                                        break
                                    print('video Date: ',video_Date)
                                    time.sleep(0.5)
                                    for channel_subscribers in browser.find_elements_by_xpath('//*[@id="owner-sub-count"]'):
                                        channel_subscribers = channel_subscribers.get_attribute('innerText').replace('subscribers', '').strip()
                                        break
                                    print('Channel Subscribers: ',channel_subscribers)
                                    time.sleep(0.5)
                                    for channel_name in browser.find_elements_by_xpath('//*[@id="text-container"]'):
                                        channel_name = channel_name.get_attribute('innerText').strip()
                                        break
                                    print('Channel Name: ',channel_name)
                                    time.sleep(0.5)
                                    for hash_tag in browser.find_elements_by_xpath('//*[@class="super-title style-scope ytd-video-primary-info-renderer"]'):
                                        hash_tag = hash_tag.get_attribute('innerText').strip()
                                        break
                                    # print(hash_tag)
                                    time.sleep(0.5)
                                    try:
                                        for more_button_desc in browser.find_elements_by_xpath('//*[@class="more-button style-scope ytd-video-secondary-info-renderer"]'):
                                            more_button_desc.click()
                                            time.sleep(1)
                                            break
                                    except:pass
                                    time.sleep(0.5)
                                    for desc in browser.find_elements_by_xpath('//*[@id="description"]'):
                                        desc = desc.get_attribute('innerText')
                                        break
                                    # print(desc)
                                    time.sleep(0.5)
                                    for comments in browser.find_elements_by_xpath('//*[@id="comments"]'):
                                        comments = comments.get_attribute('innerText')
                                        break
                                    # print(comments)

                                    if global_var.Search_keyword.lower() in video_Title.lower():
                                        keywordFound = True
                                    if global_var.Search_keyword.lower() in comments.lower():
                                        keywordFound = True
                                    if global_var.Search_keyword.lower() in desc.lower():
                                        keywordFound = True
                                    if global_var.Search_keyword.lower() in hash_tag.lower():
                                        keywordFound = True
                                    if global_var.Search_keyword.lower() in channel_name.lower():
                                        keywordFound = True
                                    if keywordFound == True:

                                        Search_keyword = global_var.Search_keyword

                                        sheet['A1'] = 'Date And Time'
                                        sheet['B1'] = 'Video Title'
                                        sheet['C1'] = 'Video Views'
                                        sheet['D1'] = 'Video Upload Date'
                                        sheet['E1'] = 'Cannel Subscribers'
                                        sheet['F1'] = 'Channel Name'
                                        sheet['G1'] = 'Video Link'
                                        sheet['H1'] = 'Channel Link'
                                        sheet['I1'] = 'Search Keyword'

                                        sheet['A' + str(max_row) + ''] = str(date_time)
                                        sheet['B' + str(max_row) + ''] = str(video_Title)
                                        sheet['C' + str(max_row) + ''] = str(video_Views)
                                        sheet['D' + str(max_row) + ''] = str(video_Date)
                                        sheet['E' + str(max_row) + ''] = str(channel_subscribers)
                                        sheet['F' + str(max_row) + ''] = str(channel_name)
                                        sheet['G' + str(max_row) + ''] = str(youtube_video_link)
                                        sheet['H' + str(max_row) + ''] = str(youtube_channel_link)
                                        sheet['I' + str(max_row) + ''] = str(Search_keyword)

                                        book.save(
                                            "D:\\Youtube Work\\Youtube Video Details.xlsx")  # save the file after write data
                                        print('Insert on xlsx Sheet')
                                        data_count += 1
                                        max_row += 1
                                        print(str(data_count) + ' Youtube Videos Mila Hai\n')
                                        time.sleep(0.5)
                                    else:pass
                                    browser.close()
                                    browser.switch_to.window(browser.window_handles[0])
                                    time.sleep(0.5)
                                else:pass
                            else:pass
                            b = False
                        except Exception as e:
                            exc_type, exc_obj, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            print("Error ON : ", sys._getframe().f_code.co_name + "--> " + str(e), "\n", exc_type, "\n",
                                  fname,
                                  "\n", exc_tb.tb_lineno)
                            time.sleep(5)
                            b = True
                No_result_found = ''
                try:
                    for No_result_found in browser.find_elements_by_xpath('//*[@id="message"]'):
                        No_result_found = No_result_found.get_attribute('innerText').strip()
                except:pass
                if 'No more results' in str(No_result_found):
                    print('No More Results Found')
                    print('!!! -_- Process Done -_- !!!')
                    print('Total Videos Found :' + str(data_count))
                    browser.close()
                    # sheet = book.create_sheet()
                    # sheet.title = str(date_time)
                    # book.save(
                    #     "D:\\Youtube Work\\Youtube Video Details.xlsx")
                    wx.MessageBox('!!! -_- Process Done -_- !!!  Total Videos Found : ' + str(data_count), 'Youtube', wx.OK | wx.ICON_INFORMATION)
                    # sys.exit()
                browser.execute_script("window.scrollTo(0, "+str(y)+")")
                y += 1080
                a = True
            else:
                print('!!! -_- Process Done -_- !!!')
                print('Total Videos Found :' + str(data_count))
                # sheet = book.create_sheet()
                # sheet.title = str(date_time)
                # book.save(
                #     "D:\\Youtube Work\\Youtube Video Details.xlsx")
                browser.close()
                wx.MessageBox('!!! -_- Process Done -_- !!!  Total Videos Found : ' + str(data_count), 'Youtube', wx.OK | wx.ICON_INFORMATION)
                # sys.exit()
                a = False
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print("Error ON : ", sys._getframe().f_code.co_name + "--> " + str(e), "\n", exc_type, "\n", fname,
                  "\n", exc_tb.tb_lineno)
            a = True
            time.sleep(5)


# def Local_Connection_Function():
#     a = 0
#     while a == 0:
#         try:
#             mydb_Local = mysql.connector.connect(
#                 host='127.0.0.1',
#                 user='root',
#                 passwd='Premjogu007',
#                 database='youtube')
#             print('SQL Local Connection Connected')
#             return mydb_Local
#         except mysql.connector.ProgrammingError as e:
#             exc_type, exc_obj, exc_tb = sys.exc_info()
#             fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
#             Error_text = "Error ON : " + " " + str(sys._getframe().f_code.co_name) + " ---> " + str(e) + " , " + str(
#                 exc_type) + " , " + str(fname) + "<br> Error Line Number: " + str(exc_tb.tb_lineno)
#             print(Error_text)
#             a = 0
#             time.sleep(5)


# def data_insert(video_Title,video_Views,video_Date,channel_subscribers,channel_name,youtube_video_link,youtube_channel_link,Search_keyword,data_count):
#     mydb_Local = Local_Connection_Function()
#     mycursorLocal = mydb_Local.cursor()
#     c = True
#     while c == True:
#         try:
#             Duplicate_Email = "Select * from youtube_scraper where video_url = '" + str(youtube_video_link) + "'"
#             mycursorLocal.execute(Duplicate_Email)
#             results = mycursorLocal.fetchall()
#             if len(results) > 0:
#                 print('Duplicate video')
#                 c = False
#             else:
#                 insert_data = "INSERT INTO youtube_scraper(video_title,video_url,video_view,video_upload_date,channel_name,channel_url,channel_subscriber,found_keyword)VALUES(%s,%s,%s,%s,%s,%s,%s,%s)"
#                 values = (str(video_Title), str(youtube_video_link), str(video_Views), str(video_Date), str(channel_name),
#                           str(youtube_channel_link), str(channel_subscribers), str(Search_keyword))
#                 mycursorLocal.execute(insert_data, values)
#                 mydb_Local.commit()
#                 print('Insert')
#                 c = False
#         except Exception as e:
#             exc_type, exc_obj, exc_tb = sys.exc_info()
#             fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
#             print("Error ON : ", sys._getframe().f_code.co_name + "--> " + str(e), "\n", exc_type, "\n", fname,
#                   "\n", exc_tb.tb_lineno)
#             c = True
#             time.sleep(3)

# def Excel_Insert(video_Title,video_Views,video_Date,channel_subscribers,channel_name,youtube_video_link,youtube_channel_link,Search_keyword,row_count):
#     book = Workbook()
#     sheet = book.active
#
#     sheet['A1'] = 'Video Title'
#     sheet['B1'] = 'Video Views'
#     sheet['C1'] = 'Video Upload Date'
#     sheet['D1'] = 'Cannel Subscribers'
#     sheet['E1'] = 'Channel Name'
#     sheet['F1'] = 'Video Link'
#     sheet['G1'] = 'Channel Link'
#     sheet['H1'] = 'Search Keyword'
#
#     sheet['A'+str(row_count)+''] = str(video_Title)
#     sheet['B'+str(row_count)+''] = str(video_Views)
#     sheet['C'+str(row_count)+''] = str(video_Date)
#     sheet['D'+str(row_count)+''] = str(channel_subscribers)
#     sheet['E'+str(row_count)+''] = str(channel_name)
#     sheet['F'+str(row_count)+''] = str(youtube_video_link)
#     sheet['G'+str(row_count)+''] = str(youtube_channel_link)
#     sheet['H'+str(row_count)+''] = str(Search_keyword)
#
#     book.save("D:\\Youtube Work\\Youtube Video Details.xlsx")  # save the file after write data
#     print('Insert on xlsx Sheet')







